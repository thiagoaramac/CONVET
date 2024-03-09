import pandas as pd
import os
import numpy as np
import openpyxl
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName
import xlwings as xw
import win32com
from win32com.client import Dispatch


def compilar_notas():
    # Dados iniciais ---------------------------------------------------------------------------------------------------
    print('------------------------------------------------')
    print('Dados Iniciais:')
    print('')
    print('Eixos temáticos:')
    eixos_tematicos = {
        "Língua Portuguesa": range(1, 11),
        "Direito administrativo e constitucional": range(11, 16),
        "Informática": range(16, 21),
        "Conhecimentos Específicos": range(21, 61)
    }
    eixos_tematicos_values = list(eixos_tematicos)
    for i in range(len(eixos_tematicos)):
        print('   ->' + eixos_tematicos_values[i])
    print('')
    q_basico_csv = 0.17  # Quanto vale uma questão básica no .CSV?
    q_especifica_csv = 0.17  # Quanto vale uma questão específica no .CSV?

    print('As questões básicas no CSV valem: ' + str(q_basico_csv))
    print('As questões específicas no CSV valem: ' + str(q_especifica_csv))
    print('')

    print('Pesos das provas:')
    peso_basico = 5

    peso_especifico1 = 5
    peso_especifico2 = 5
    peso_especifico3 = 5
    peso_especifico4 = 5
    pesos_especificos = [peso_especifico1, peso_especifico2, peso_especifico3, peso_especifico4]

    print('   ->Questões da prova básica: ' + str(peso_basico))
    for i in range(len(eixos_tematicos)):
        print(f'   ->Questões do Eixo Temático {i + 1}: ' + str(pesos_especificos[i]))
    print('------------------------------------------------')

    # ------------------------------------------------------------------------------------------------------------------

    input_files_path = os.getcwd() + '\\input-files\\'
    files = os.listdir(input_files_path)

    input_especifico = [file for file in files if "GABAR" in file][0]

    # Gera o CSV_Especifico --------------------------------------------------------------------------------------------
    df_especifico = pd.read_csv(os.getcwd() + '\\input-files\\' + input_especifico)
    df_especifico.pop('Estado')
    df_especifico.pop('Iniciado em')
    df_especifico.pop('Completo')
    df_especifico.pop('Tempo utilizado')
    df_especifico.pop('Avaliar/10,00')
    df_especifico = df_especifico[~df_especifico['Sobrenome'].str.startswith('Média geral')]
    df_especifico['Aluno'] = (df_especifico['Nome'] + ' ' + df_especifico['Sobrenome']).str.title()
    df_especifico = df_especifico.drop(['Nome', 'Sobrenome'], axis=1)
    df_especifico = df_especifico.replace('-', '0,00', regex=True)
    df_especifico = df_especifico.replace(',', '.', regex=True)
    df_especifico = df_especifico.apply(pd.to_numeric, errors='ignore')
    df_especifico = df_especifico.replace(0.17, 1/6, regex=True)

    df_especifico = df_especifico[['Aluno'] + [col for col in df_especifico.columns if col != 'Aluno']]
    df_especifico.to_csv(os.getcwd() + '\\output-files\\CSV_Especifico.csv')

    # Cria a coluna com a soma de todas as questões --------------------------------------------------------------------
    multiplicador = 1
    try:
        df_especifico['Língua Portuguesa'] = df_especifico[
            ['Q. 1 /0,20', 'Q. 2 /0,20', 'Q. 3 /0,20', 'Q. 4 /0,20', 'Q. 5 /0,20',
             'Q. 6 /0,20', 'Q. 7 /0,20', 'Q. 8 /0,20', 'Q. 9 /0,20', 'Q. 10 /0,20']
        ].sum(axis=1)

        df_especifico['Direito administrativo e constitucional'] = df_especifico[
            ['Q. 11 /0,20', 'Q. 12 /0,20', 'Q. 13 /0,20', 'Q. 14 /0,20', 'Q. 15 /0,20']
        ].sum(axis=1)

        df_especifico['Informática'] = df_especifico[
            ['Q. 16 /0,20', 'Q. 17 /0,20', 'Q. 18 /0,20', 'Q. 19 /0,20', 'Q. 20 /0,20']
        ].sum(axis=1)

        df_especifico['Prova Objetiva Básica'] = df_especifico[
            ['Língua Portuguesa', 'Direito administrativo e constitucional', 'Informática']
        ].sum(axis=1)
        multiplicador = 5
    except:
        pass
    try:
        df_especifico['Língua Portuguesa'] = df_especifico[
            ['Q. 1 /0,17', 'Q. 2 /0,17', 'Q. 3 /0,17', 'Q. 4 /0,17', 'Q. 5 /0,17',
             'Q. 6 /0,17', 'Q. 7 /0,17', 'Q. 8 /0,17', 'Q. 9 /0,17', 'Q. 10 /0,17']
        ].sum(axis=1)

        df_especifico['Direito administrativo e constitucional'] = df_especifico[
            ['Q. 11 /0,17', 'Q. 12 /0,17', 'Q. 13 /0,17', 'Q. 14 /0,17', 'Q. 15 /0,17']
        ].sum(axis=1)

        df_especifico['Informática'] = df_especifico[
            ['Q. 16 /0,17', 'Q. 17 /0,17', 'Q. 18 /0,17', 'Q. 19 /0,17', 'Q. 20 /0,17']
        ].sum(axis=1)

        df_especifico['Prova Objetiva Básica'] = df_especifico[
            ['Língua Portuguesa', 'Direito administrativo e constitucional', 'Informática']
        ].sum(axis=1)
        multiplicador = 6
    except:
        pass

    df_especifico.to_csv(os.getcwd() + '\\output-files\\CSV_Especificoteste.csv')

    try:
        df_especifico['Prova Objetiva Específica'] = df_especifico[
            ['Q. 21 /0,17', 'Q. 22 /0,17', 'Q. 23 /0,17', 'Q. 24 /0,17', 'Q. 25 /0,17',
             'Q. 26 /0,17', 'Q. 27 /0,17', 'Q. 28 /0,17', 'Q. 29 /0,17', 'Q. 30 /0,17',
             'Q. 31 /0,17', 'Q. 32 /0,17', 'Q. 33 /0,17', 'Q. 34 /0,17', 'Q. 35 /0,17',
             'Q. 36 /0,17', 'Q. 37 /0,17', 'Q. 38 /0,17', 'Q. 39 /0,17', 'Q. 40 /0,17',
             'Q. 41 /0,17', 'Q. 42 /0,17', 'Q. 43 /0,17', 'Q. 44 /0,17', 'Q. 45 /0,17',
             'Q. 46 /0,17', 'Q. 47 /0,17', 'Q. 48 /0,17', 'Q. 49 /0,17', 'Q. 50 /0,17',
             'Q. 51 /0,17', 'Q. 52 /0,17', 'Q. 53 /0,17', 'Q. 54 /0,17', 'Q. 55 /0,17',
             'Q. 56 /0,17', 'Q. 57 /0,17', 'Q. 58 /0,17', 'Q. 59 /0,17', 'Q. 60 /0,17']
        ].sum(axis=1)
    except:
        pass
    try:
        df_especifico['Prova Objetiva Específica'] = df_especifico[
            ['Q. 21 /0,20', 'Q. 22 /0,20', 'Q. 23 /0,20', 'Q. 24 /0,20', 'Q. 25 /0,20',
             'Q. 26 /0,20', 'Q. 27 /0,20', 'Q. 28 /0,20', 'Q. 29 /0,20', 'Q. 30 /0,20',
             'Q. 31 /0,20', 'Q. 32 /0,20', 'Q. 33 /0,20', 'Q. 34 /0,20', 'Q. 35 /0,20',
             'Q. 36 /0,20', 'Q. 37 /0,20', 'Q. 38 /0,20', 'Q. 39 /0,20', 'Q. 40 /0,20',
             'Q. 41 /0,20', 'Q. 42 /0,20', 'Q. 43 /0,20', 'Q. 44 /0,20', 'Q. 45 /0,20',
             'Q. 46 /0,20', 'Q. 47 /0,20', 'Q. 48 /0,20', 'Q. 49 /0,20', 'Q. 50 /0,20']
        ].sum(axis=1)
    except:
        pass

    for i in range(61):
        try:
            df_especifico.pop('Q. ' + str(i) + ' /0,20')
        except:
            pass
    for i in range(61):
        try:
            df_especifico.pop('Q. ' + str(i) + ' /0,17')
        except:
            pass

    df_especifico["Prova Objetiva Específica"] = df_especifico["Prova Objetiva Específica"].apply(lambda x: round(x, 2))
    df_especifico = df_especifico

    # ------------------------------------------------------------------------------------------------------------------
    df_especifico.reset_index(drop=True, inplace=True)
    df_especifico.to_csv(os.getcwd() + '\\output-files\\CSV_Especifico.csv')
    print("Arquivo CSV_Especifico.csv gerado com sucesso!")
    # Cria o CSV_NotasFinais -------------------------------------------------------------------------------------------
    df_nota_final = df_especifico

    df_nota_final['Prova Objetiva (Geral + Específica)'] = df_nota_final[
        ['Prova Objetiva Específica',
        'Prova Objetiva Básica']
        ].sum(axis=1).round()

    df_nota_final['Nota Final'] = df_nota_final[['Prova Objetiva (Geral + Específica)']].round(2)

    df_nota_final = df_nota_final.sort_values(by='Prova Objetiva Específica', ascending=False)
    df_nota_final = df_nota_final.drop_duplicates(subset='Aluno', keep='first')
    df_nota_final = df_nota_final.fillna(0)
    df_nota_final = df_nota_final.sort_values(by='Aluno', ascending=True)
    df_nota_final.reset_index(drop=True, inplace=True)

    df_nota_final["Língua Portuguesa"] = df_nota_final["Língua Portuguesa"] * multiplicador
    df_nota_final["Direito administrativo e constitucional"] = df_nota_final["Direito administrativo e constitucional"] * multiplicador
    df_nota_final["Informática"] = df_nota_final["Informática"] * multiplicador
    df_nota_final["Prova Objetiva Básica"] = df_nota_final["Prova Objetiva Básica"] * multiplicador
    df_nota_final["Prova Objetiva Específica"] = df_nota_final["Prova Objetiva Específica"] * multiplicador
    df_nota_final["Prova Objetiva (Geral + Específica)"] = df_nota_final["Prova Objetiva (Geral + Específica)"] * multiplicador
    df_nota_final["Nota Final"] = df_nota_final["Nota Final"] * multiplicador

    df_nota_final.to_csv(os.getcwd() + '\\output-files\\CSV_NotasFinais.csv')
    print("Arquivo CSV_NotasFinais.csv gerado com sucesso!")
    # Limpa a memória RAM
    del df_especifico, df_nota_final

def rankear_alunos(simulado_num):
    # Dados iniciais ---------------------------------------------------------------------------------------------------
    nota_maxima_basica = 20
    nota_maxima_especifica = 30

    email_database = 'Usuários (1).xlsx'  # O arquivo deve estar na pasta '\list-email'

    print('------------------------------------------------')
    print('Dados Iniciais:')
    print('')
    print('Nota máxima da prova básica: ' + str(nota_maxima_basica))
    print('Nota máxima da prova específica: ' + str(nota_maxima_especifica))
    print('')
    print('------------------------------------------------')
    nota_maxima = nota_maxima_basica + nota_maxima_especifica
    # ------------------------------------------------------------------------------------------------------------------

    input_files_path = os.getcwd() + '\\output-files\\'
    files = os.listdir(input_files_path)

    input_ranking = [file for file in files if "CSV_NotasFinais" in file][0]

    df_rank = pd.read_csv(os.getcwd() + '\\output-files\\' + input_ranking)

    notas = df_rank.columns[2:].tolist()
    for nota in notas:
        df_rank = df_rank.sort_values(by=[nota], ascending=False)
        df_rank['Ranking ' + nota] = df_rank[nota].rank(ascending=False, method='dense').astype(int)
        df_rank = df_rank.sort_index()

    df_rank = df_rank.sort_values(by='Ranking Nota Final', ascending=True)
    df_rank.reset_index(drop=True, inplace=True)
    df_rank.to_csv(os.getcwd() + '\\output-files\\CSV_Ranking.csv')

    df_rank['Ranking Língua Portuguesa'] = (df_rank['Ranking Língua Portuguesa'].astype(str))
    df_rank['Ranking Direito administrativo e constitucional'] = (
            df_rank['Ranking Direito administrativo e constitucional'].astype(str))
    df_rank['Ranking Informática'] = (df_rank['Ranking Informática'].astype(str))
    df_rank['Ranking Prova Objetiva Básica'] = (df_rank['Ranking Prova Objetiva Básica'].astype(str))
    df_rank['Ranking Prova Objetiva Específica'] = (df_rank['Ranking Prova Objetiva Específica'].astype(str))
    df_rank['Ranking Prova Objetiva (Geral + Específica)'] = (
            df_rank['Ranking Prova Objetiva (Geral + Específica)'].astype(str))
    df_rank['Ranking Nota Final'] = (df_rank['Ranking Nota Final'].astype(str))

    df_rank['Texto1'] = "Oi, " + df_rank['Aluno'].str.split().str[0] + "! Tudo bem?\n"
    df_rank['Texto2'] = ("Sua Nota Final: " + df_rank['Nota Final'].astype(str) + f"/{str(nota_maxima)}\n")
    df_rank['Texto4'] = ( "Seu ranking da prova Objetiva Básica: " + df_rank['Ranking Prova Objetiva Básica'].astype(str) + "ª nota mais alta\n")
    df_rank['Texto5'] = ("Seu ranking da prova Objetiva Específica: " + df_rank['Ranking Prova Objetiva Específica']. astype(str) + "ª nota mais alta\n")
    df_rank['Texto6'] = ("Seu ranking geral: " + df_rank['Ranking Nota Final'].astype(str) + "ª nota mais alta\n\n")
    df_rank['Texto7'] = "Nas disciplinas de conhecimentos básicos:\n"
    df_rank['Texto8'] = ("Língua Portuguesa: " + df_rank['Ranking Língua Portuguesa'].astype(str) + "ª nota mais alta\n")
    df_rank['Texto9'] = ("Direito administrativo e constitucional: " + df_rank[ 'Ranking Direito administrativo e constitucional'].astype(str) + "ª nota mais alta\n")
    df_rank['Texto10'] = ("Informática: " + df_rank['Ranking Informática'].astype(str) + "ª nota mais alta\n")
    df_rank['Texto15'] = ("Esse simulado teve " + str(len(df_rank)) + " alunos\nContinue firme e bons estudos!!")

    df_rank['Feedback'] = df_rank[
        ['Texto1', 'Texto2', 'Texto4', 'Texto5', 'Texto6', 'Texto7', 'Texto8', 'Texto9', 'Texto10', 'Texto15']].apply(lambda x: '\n'.join(x), axis=1)
    df_rank = df_rank.drop(
        columns=['Texto1', 'Texto2', 'Texto4', 'Texto5', 'Texto6', 'Texto7', 'Texto8', 'Texto9', 'Texto10', 'Texto15'])

    df_rank['Feedback'] = df_rank['Feedback'].apply(lambda x: '\n'.join(line for line in x.splitlines() if line.strip()))
    df_rank['Feedback'] = df_rank['Feedback'].apply(lambda x: '\n'.join(line if line != 'Nas disciplinas:' else '\n' + line for line in x.splitlines() if line.strip()))
    df_rank['Feedback'] = df_rank['Feedback'].str.replace(' -ª nota mais alta', ' -')

    df_rank = df_rank[df_rank['Ranking Nota Final'] != "-"]
    df_rank.pop('Unnamed: 0')

    df_rank.reset_index(drop=True, inplace=True)

    df_rank.to_csv(os.getcwd() + '\\output-files\\CSV_Ranking.csv')
    print("Arquivo CSV_Ranking.csv gerado com sucesso!")

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    email_feedback = []
    for i in range(len(df_rank)):
        var_aluno = df_rank.iat[i, 0].split()[0]  # Aluno
        var1 = df_rank.iat[i, 1].astype(str)      # Nota Língua Portuguesa
        var2 = df_rank.iat[i, 2].astype(str)      # Nota Direito administrativo e constitucional
        var3 = df_rank.iat[i, 3].astype(str)      # Nota Informática
        var7 = df_rank.iat[i, 4].astype(str)      # Nota Prova Objetiva Básica
        var8 = df_rank.iat[i, 5].astype(str)      # Nota Prova Objetiva Específica
        var10 = df_rank.iat[i, 6].astype(str)     # Nota Prova Objetiva (Geral + Específica)
        var11 = df_rank.iat[i, 7].astype(str)     # Nota Final
        var12 = df_rank.iat[i, 8]                 # Ranking Língua Portuguesa
        var13 = df_rank.iat[i, 9]                 # Ranking Direito administrativo e constitucional
        var14 = df_rank.iat[i, 10]                # Ranking Informática
        var18 = df_rank.iat[i, 11]                # Ranking Prova Objetiva Básica
        var19 = df_rank.iat[i, 12]                # Ranking Prova Objetiva Específica
        var21 = df_rank.iat[i, 13]                # Ranking Prova Objetiva (Geral + Específica)
        var22 = df_rank.iat[i, 14]                # Ranking Final

        filtered_values = df_rank[df_rank['Prova Objetiva Básica'] != 0]['Prova Objetiva Básica']
        var23 = round(filtered_values.mean(), 2)  # Média conhecimentos básicos

        filtered_values = df_rank[df_rank['Prova Objetiva Específica'] != 0]['Prova Objetiva Específica']
        var24 = round(filtered_values.mean(), 2)  # Média conhecimentos específicos

        filtered_values = df_rank[df_rank['Nota Final'] != 0]['Nota Final']
        var26 = round(filtered_values.mean(), 2)  # Média Nota Final

        html_string1 = f"<p>Oi, {var_aluno}! Tudo bem? &#128516;<br/></p>"

        if len(str(nota_maxima_basica)) == 2:
            nota_maxima_basica = str(nota_maxima_basica) + '.00'
        if len(str(nota_maxima_especifica)) == 2:
            nota_maxima_especifica = str(nota_maxima_especifica) + '.00'
        if len(str(nota_maxima)) == 2:
            nota_maxima = str(nota_maxima) + '.00'

        html_string1a = "<p><strong>Suas notas:</strong><br/>"
        html_string1b = f"Prova objetiva - Conhecimentos Gerais: <strong>{var7}0/{nota_maxima_basica}</strong><br/>"
        html_string1c = f"Prova objetiva - Conhecimentos Específicos: <strong>{var8}0/{nota_maxima_especifica}</strong><br/>"
        html_string3 = f"Sua Nota Final: <strong>{var11}0/{str(nota_maxima)}</strong><br/>"

        html_string2 = "<p><strong>Seu feedback na prova como um todo:</strong><br/>"

        if var18 == '-':
            html_string5 = f"Seu ranking da prova Objetiva - Conhecimentos Gerais: <strong>{var18}</strong><br/>"
        if var18 != '-':
            html_string5 = f"Seu ranking da prova Objetiva - Conhecimentos Gerais: <strong>{var18}ª</strong> nota mais alta<br/>"

        if var19 == '-':
            html_string6 = f"Seu ranking da prova Objetiva - Conhecimentos Específicos: <strong>{var19}</strong><br/>"
        if var19 != '-':
            html_string6 = f"Seu ranking da prova Objetiva - Conhecimentos Específicos: <strong>{var19}ª</strong> nota mais alta<br/>"

        if var22 == '-':
            html_string7 = f"Seu ranking geral: <strong>{var22}</strong></p>"
        if var22 != '-':
            html_string7 = f"Seu ranking geral: <strong>{var22}ª</strong> nota mais alta</p>"

        html_string8 = "<p><strong>Seu feedback por área de conhecimento:</strong><br/>"

        if var12 == '-':
            html_string9 = f"Língua Portuguesa: <strong>{var12}</strong><br/>"
        if var12 != '-':
            html_string9 = f"Língua Portuguesa: <strong>{var12}ª</strong> nota mais alta<br/>"

        if var13 == '-':
            html_string10 = f"Direito administrativo e constitucional: <strong>{var13}</strong><br/>"
        if var13 != '-':
            html_string10 = f"Direito administrativo e constitucional: <strong>{var13}ª</strong> nota mais alta<br/>"

        if var14 == '-':
            html_string11 = f"Informática: <strong>{var14}</strong><br/>"
        if var14 != '-':
            html_string11 = f"Informática: <strong>{var14}ª</strong> nota mais alta<br/>"

        html_string14 = f"<p><strong>Esse simulado foi feito por <u>{str(len(df_rank))}</u> alunos</strong><br/>"
        html_string15 = f"Média geral (Conhecimentos Gerais): <strong>{var23}/{str(nota_maxima_basica)}</strong><br/>"
        html_string16 = f"Média geral (Conhecimentos Específicos): <strong>{var24}/{str(nota_maxima_especifica)}</strong><br/>"
        html_string17b = f"Média geral do simulado: <strong>{var26}/{str(nota_maxima)}</strong><br/><br/>"
        html_string18 = "Bons estudos!!</p>"
        html_string19 = '<p><font color="gray"><b><i><h1 style="font-size:8pt; ">'
        html_string20 = "Lembrando que esse é um email automático do CONVET!"
        html_string21 = f'</h1></i></b><i><h1 style="font-size:8pt; ">'
        html_string22 = "Se você tentar respondê-lo, ninguém vai ver &#128542;"
        html_string23 = "</h1></i></font></p>"
        html_string24 = f'<img src="https://www.concursosconvet.com.br/assets/img/logotipo/logotipo.png"  width="100" height="100">'

        html_strings = [html_string1, html_string1a, html_string1b, html_string1c, html_string3, html_string2,
                        html_string5, html_string6, html_string7, html_string8, html_string9, html_string10,
                        html_string11, html_string14, html_string15, html_string16, html_string17b, html_string18,
                        html_string19, html_string20, html_string21, html_string22, html_string23, html_string24]

        html_string = "".join(html_strings)
        email_feedback.append(html_string)

    df_rank['Email Feedback'] = email_feedback
    df_email_list = pd.read_excel(os.getcwd() + '\\list-email\\' + email_database)

    df_email_list['Aluno'] = (df_email_list['firstname'] + ' ' + df_email_list['lastname']).str.title()

    try:
        df_email_list.pop('id')
        df_email_list.pop('username')
        df_email_list.pop('firstname')
        df_email_list.pop('lastname')
        df_email_list.pop('idnumber')
        df_email_list.pop('institution')
        df_email_list.pop('department')
        df_email_list.pop('phone1')
        df_email_list.pop('phone2')
        df_email_list.pop('city')
        df_email_list.pop('url')
        df_email_list.pop('icq')
        df_email_list.pop('skype')
        df_email_list.pop('aim')
        df_email_list.pop('yahoo')
        df_email_list.pop('msn')
        df_email_list.pop('country')
        df_email_list.pop('profile_field_cpf')
    except Exception as e:
        print(f"Aconteceu um erro: {e}")

    df_email_list = df_email_list[['Aluno', 'email']]

    df_email_list['Aluno'] = df_email_list['Aluno'].str.replace('  ', ' ')
    df_rank['Aluno'] = df_rank['Aluno'].str.replace('  ', ' ')

    df_email_feedback = pd.merge(df_rank, df_email_list, on='Aluno', how='left')

    try:
        df_email_feedback.pop('Língua Portuguesa')
        df_email_feedback.pop('Direito administrativo e constitucional')
        df_email_feedback.pop('Informática')
        df_email_feedback.pop('Prova Objetiva Específica')
        df_email_feedback.pop('Prova Objetiva Básica')
        df_email_feedback.pop('Prova Objetiva (Geral + Específica)')
        df_email_feedback.pop('Nota Final')

        df_email_feedback.pop('Ranking Língua Portuguesa')
        df_email_feedback.pop('Ranking Direito administrativo e constitucional')
        df_email_feedback.pop('Ranking Informática')
        df_email_feedback.pop('Ranking Prova Objetiva Específica')
        df_email_feedback.pop('Ranking Prova Objetiva Básica')
        df_email_feedback.pop('Ranking Prova Objetiva (Geral + Específica)')
        df_email_feedback.pop('Ranking Nota Final')
        df_email_feedback.pop('Feedback')
    except Exception as e:
        print(e)

    df_email_feedback.to_excel(os.getcwd() + f'\\list-email\\Excel_Email_List - Simulado CIDASC {simulado_num}.xlsx',
                               index=False)

    # try:
    #    df_feedback = df_rank[['Aluno', 'Email Feedback']].copy()
    #    df_feedback.to_excel(os.getcwd() + '\\list-email\\Excel_Email_Feedback.xlsx', index = False)
    #    df_rank = df_rank.drop(columns = ['Email Feedback'])
    #    print("Arquivo Excel_Email_Feedback.xlsx gerado com sucesso!")
    # except Exception as e:
    #    print(f"Aconteceu um erro: {e}")

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------

    try:
        df_rank.to_excel(os.getcwd() + f'\\output-files\\Excel_Ranking - Simulado CIDASC {simulado_num}.xlsx', index=False)
        print('------------------------------------------------')
        print(f'Arquivo Excel_Ranking - Simulado CIDASC {simulado_num}.xlsx gerado com sucesso!')
        print('')
    except Exception as e:
        print(f'Aconteceu um erro: {e}')


def arrumar_excel(simulado_num):
    input_files_path = os.getcwd() + f'\\output-files\\Excel_Ranking - Simulado CIDASC {simulado_num}.xlsx'
    output_files_path = os.getcwd() + f'\\output-files\\Excel_Ranking - Simulado CIDASC {simulado_num}.xlsx'

    try:
        workbook = openpyxl.load_workbook(input_files_path)
        sheet = workbook.active

        # 1. Autoadjust column 'A' width
        sheet.column_dimensions['A'].auto_size = True

        # 2. Set cells 'B1' to 'U1' to Wrap Text
        for column in range(ord('B'), ord('U') + 1):
            col_letter = chr(column)
            sheet[f'{col_letter}1'].alignment = Alignment(wrap_text=True)

        # 3. Set cells 'B1' to 'U1' to width=16
        for column in range(ord('B'), ord('U') + 1):
            col_letter = chr(column)
            sheet.column_dimensions[col_letter].width = 16

        # 4. Set column 'V' to width=64
        sheet.column_dimensions['V'].width = 64

        # 5. Set column 'V' to Wrap Text
        sheet['V1'].alignment = Alignment(wrap_text=True)

        # 6. Set row 2 to the last one to center text in cells
        # 7. Set row 2 to the last one to center align in cells
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

        # 8. Set column 'V' to left align the text
        for cell in sheet['V']:
            cell.alignment = Alignment(horizontal="left")

        # 9. Set row 1 height=60
        sheet.row_dimensions[1].height = 60

        # 10. Set row 2 to the last one height=225
        for row in range(2, sheet.max_row + 1):
            sheet.row_dimensions[row].height = 225

        # 11. Vertical align all cells in row 1 to the bottom
        sheet.column_dimensions['A'].width = 42
        for cell in sheet[1]:
            cell.alignment = Alignment(vertical="bottom")

        # 12. Set all cells to Wrap Text
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

        # 14. Put macro
        vba_code = ("""
        Option Explicit

        Private Declare PtrSafe Function OpenClipboard Lib "user32" (ByVal hwnd As LongPtr) As Long
        Private Declare PtrSafe Function CloseClipboard Lib "user32" () As Long
        Private Declare PtrSafe Function EmptyClipboard Lib "user32" () As Long
        Private Declare PtrSafe Function SetClipboardData Lib "user32" (ByVal wFormat As Long, ByVal hMem As LongPtr) As LongPtr
        Private Declare PtrSafe Function GlobalAlloc Lib "kernel32" (ByVal wFlags As Long, ByVal dwBytes As LongPtr) As LongPtr
        Private Declare PtrSafe Function GlobalLock Lib "kernel32" (ByVal hMem As LongPtr) As LongPtr
        Private Declare PtrSafe Function GlobalUnlock Lib "kernel32" (ByVal hMem As LongPtr) As Long
        Private Declare PtrSafe Function GlobalFree Lib "kernel32" (ByVal hMem As LongPtr) As LongPtr
        Private Declare PtrSafe Function lstrcpy Lib "kernel32" (ByVal lpStr1 As Any, ByVal lpStr2 As Any) As LongPtr
        
        Private Const CF_TEXT = 1&
        Private Const GMEM_MOVEABLE = &H2
        
        Sub CopyContent()
        Call StringToClipboard(ActiveCell.Value)
        End Sub
        
        Private Sub StringToClipboard(strText As String)
        Dim lngIdentifier As LongPtr, lngPointer As LongPtr
        lngIdentifier = GlobalAlloc(GMEM_MOVEABLE, Len(strText) + 1)
        lngPointer = GlobalLock(lngIdentifier)
        Call lstrcpy(ByVal lngPointer, strText)
        Call GlobalUnlock(lngIdentifier)
        Call OpenClipboard(0&)
        Call EmptyClipboard
        Call SetClipboardData(CF_TEXT, lngIdentifier)
        Call CloseClipboard
        Call GlobalFree(lngIdentifier)
        End Sub

        """
                    )
        # app = xw.App(visible = True, add_book = False)
        # wb = app.books.open(input_files_path)

        # wb.vba.add_module(vba_code, "MyModule")
        # wb.api.Application.OnKey("^q", "MyModule.MyMacro")
        # wb.save()

        # Save the workbook
        max_row = sheet.max_row
        for row in range(1, max_row + 1):
            cell_value = sheet.cell(row=row, column=22).value  # Column 'V' is the 22nd column
            if cell_value is not None and isinstance(cell_value, str):
                sheet.cell(row=row, column=22, value=cell_value.replace('"', ''))

        workbook.save(output_files_path)
        # app.quit()
    except Exception as e:
        print(f"Aconteceu um erro: {e}")


def formatar_excel(simulado_num):
    input_files_path = os.getcwd() + f'\\output-files\\Excel_Ranking - Simulado CIDASC {simulado_num}.xlsx'
    wb = load_workbook(input_files_path)
    ws = wb['Sheet1']

    ws.column_dimensions['V'].width = 63

    for letter in ['A']:
        max_width = 0

        for row_number in range(1, ws.max_row + 1):
            if len(ws[f'{letter}{row_number}'].value) > max_width:
                max_width = len(ws[f'{letter}{row_number}'].value)
            ws['V' + str(row_number)].alignment = Alignment(wrapText=True)
            ws['A' + str(row_number)].alignment = Alignment(vertical='center', horizontal='left')

        ws.column_dimensions[letter].width = max_width + 1

    for letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U']:
        for row_number in range(1, ws.max_row + 1):
            ws[str(letter) + str(row_number)].alignment = Alignment(vertical='center', horizontal='center')

    wb.save(input_files_path)
    print('------------------------------------------------')
    print("Planilha pronta!")
    print('------------------------------------------------')
    # print("Abrindo Excel")
    # app = xw.App(visible = True, add_book = False)
    # wb = app.books.open(input_files_path)


def colocar_macro(simulado_num):
    input_files_path = os.getcwd() + f'\\output-files\\Excel_Ranking - Simulado CIDASC {simulado_num}.xlsx'

    vba_code = ("""
            Option Explicit

            Private Declare PtrSafe Function OpenClipboard Lib "user32" (ByVal hwnd As LongPtr) As Long
            Private Declare PtrSafe Function CloseClipboard Lib "user32" () As Long
            Private Declare PtrSafe Function EmptyClipboard Lib "user32" () As Long
            Private Declare PtrSafe Function SetClipboardData Lib "user32" (ByVal wFormat As Long, ByVal hMem As LongPtr) As LongPtr
            Private Declare PtrSafe Function GlobalAlloc Lib "kernel32" (ByVal wFlags As Long, ByVal dwBytes As LongPtr) As LongPtr
            Private Declare PtrSafe Function GlobalLock Lib "kernel32" (ByVal hMem As LongPtr) As LongPtr
            Private Declare PtrSafe Function GlobalUnlock Lib "kernel32" (ByVal hMem As LongPtr) As Long
            Private Declare PtrSafe Function GlobalFree Lib "kernel32" (ByVal hMem As LongPtr) As LongPtr
            Private Declare PtrSafe Function lstrcpy Lib "kernel32" (ByVal lpStr1 As Any, ByVal lpStr2 As Any) As LongPtr

            Private Const CF_TEXT = 1&
            Private Const GMEM_MOVEABLE = &H2

            Sub CopyContent()
            Call StringToClipboard(ActiveCell.Value)
            End Sub

            Private Sub StringToClipboard(strText As String)
            Dim lngIdentifier As LongPtr, lngPointer As LongPtr
            lngIdentifier = GlobalAlloc(GMEM_MOVEABLE, Len(strText) + 1)
            lngPointer = GlobalLock(lngIdentifier)
            Call lstrcpy(ByVal lngPointer, strText)
            Call GlobalUnlock(lngIdentifier)
            Call OpenClipboard(0&)
            Call EmptyClipboard
            Call SetClipboardData(CF_TEXT, lngIdentifier)
            Call CloseClipboard
            Call GlobalFree(lngIdentifier)
            End Sub

            """)

    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    workbook = excel.Workbooks.Open(Filename=input_files_path)
    excelModule = workbook.VBProject.VBComponents.Add(1)
    excelModule.CodeModule.AddFromString(vba_code)
    excel.Workbooks(1).Close(SaveChanges=1)
    excel.Application.Quit()
    del excel

    print(f'Código VBA inserido com sucesso')
