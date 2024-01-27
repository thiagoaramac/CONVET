import pandas as pd
import os
import openpyxl
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName
import xlwings as xw
import win32com
from win32com.client import Dispatch


def compilar_notas():
    input_files_path = os.getcwd() + '\\input-files\\'
    files = os.listdir(input_files_path)

    input_basico = [file for file in files if "SIC" in file][0]
    input_especifico = [file for file in files if "FIC" in file][0]
    input_discursiva = [file for file in files if "RSI" in file][0]

    # Gera o CSV_Basico ------------------------------------------------------------------------------------------------
    df_basico = pd.read_csv(os.getcwd() + '\\input-files\\' + input_basico)
    df_basico.pop('Estado')
    df_basico.pop('Iniciado em')
    df_basico.pop('Completo')
    df_basico.pop('Tempo utilizado')
    df_basico.pop('Avaliar/10,00')
    df_basico = df_basico[~df_basico['Sobrenome'].str.startswith('Média geral')]
    df_basico['Aluno'] = (df_basico['Nome'] + ' ' + df_basico['Sobrenome']).str.title()
    df_basico = df_basico.drop(['Nome', 'Sobrenome'], axis=1)
    df_basico = df_basico.replace('-', '0,00', regex=True)
    df_basico = df_basico.replace(',', '.', regex=True)
    df_basico = df_basico.apply(pd.to_numeric, errors='ignore')
    df_basico = df_basico[['Aluno'] + [col for col in df_basico.columns if col != 'Aluno']]
    df_basico.reset_index(drop=True, inplace=True)

    # Lê as matérias e cria colunas com as somas de cada matéria -------------------------------------------------------
    df_materias = pd.read_csv(os.getcwd() + '\\materias.csv')
    for index, row in df_materias.iterrows():
        materia_atual = row['C1']
        if not pd.isna(materia_atual):
            coluna_inicial = int(row['C4'])
            coluna_final = int(row['C5']) + 1
            df_basico[materia_atual] = df_basico.iloc[:, coluna_inicial:coluna_final].sum(axis=1).round(2)
    df_basico = df_basico.drop(df_basico.columns[1:51], axis=1)
    # ------------------------------------------------------------------------------------------------------------------

    # Cria a soma total da prova objetiva ------------------------------------------------------------------------------
    df_basico['Prova Objetiva Geral'] = df_basico.iloc[:, 1:].sum(axis=1).round(2)

    # ------------------------------------------------------------------------------------------------------------------
    # df_basico.to_csv(os.getcwd() + '\\input-files\\CSV_Basico.csv')


    # Gera o CSV_Especifico --------------------------------------------------------------------------------------------
    df_especifico = pd.read_csv(os.getcwd() + '\\input-files\\' + input_especifico)
    df_especifico.pop('Estado')
    df_especifico.pop('Iniciado em')
    df_especifico.pop('Completo')
    df_especifico.pop('Tempo utilizado')
    df_especifico.pop('Avaliar/10,00')
    df_especifico = df_especifico[~df_especifico['Sobrenome'].str.startswith('Média geral')]
    df_especifico['Aluno'] = (df_especifico['Nome'] + ' ' + df_especifico['Sobrenome']).str.title()
    df_especifico = df_especifico.drop(['Nome', 'Sobrenome'], axis=1)
    df_especifico = df_especifico.replace('-', '0,00', regex=True)
    df_especifico = df_especifico.replace(',', '.', regex=True)
    df_especifico = df_especifico.apply(pd.to_numeric, errors='ignore')
    df_especifico = df_especifico[['Aluno'] + [col for col in df_especifico.columns if col != 'Aluno']]
    # Cria a coluna com a soma de todas as questões ----------
    df_especifico['Prova Objetiva Específica'] = df_especifico.iloc[:, 1:].sum(axis=1)
    df_especifico = df_especifico.iloc[:, [0, -1]]
    # --------------------------------------------------------
    df_especifico.reset_index(drop=True, inplace=True)
    # df_especifico.to_csv(os.getcwd() + '\\input-files\\CSV_Especifico.csv')


    # Gera o CSV_Discursiva ------------------------------------------------------------------------------------------------
    df_discursiva = pd.read_csv(os.getcwd() + '\\input-files\\' + input_discursiva)
    df_discursiva.pop('Identificador')
    df_discursiva.pop('Status')
    df_discursiva.pop('Nota máxima')
    df_discursiva.pop('Nota pode ser alterada')
    df_discursiva.pop('Última modificação (envio)')
    df_discursiva.pop('Última modificação (nota)')
    df_discursiva.pop('Comentários de feedback')
    df_discursiva = df_discursiva.dropna(subset=['Nota'])
    df_discursiva.rename(columns={'Nome completo': 'Aluno'}, inplace=True)
    df_discursiva['Aluno'] = (df_discursiva['Aluno']).str.title()
    df_discursiva = df_discursiva.replace(',', '.', regex=True)
    df_discursiva = df_discursiva.apply(pd.to_numeric, errors='ignore')
    df_discursiva = df_discursiva.loc[df_discursiva.groupby('Aluno')['Nota'].idxmax()]
    df_discursiva.reset_index(drop=True, inplace=True)
    df_discursiva.rename(columns={'Nota': 'Prova Discursiva'}, inplace=True)
    # df_discursiva.to_csv(os.getcwd() + '\\input-files\\CSV_Discursiva.csv')


    # Cria o CSV_NotasFinais
    df_nota_final = pd.merge(df_basico, df_especifico, on='Aluno', how='outer')
    df_nota_final = pd.merge(df_nota_final, df_discursiva, on='Aluno', how='outer')
    df_nota_final['Prova Objetiva (Geral + Específica)'] = df_nota_final.iloc[:, 6:8].sum(axis=1).round(2)
    df_nota_final['Nota Final'] = df_nota_final.iloc[:, 8:10].sum(axis=1).round(2)
    df_nota_final = df_nota_final.sort_values(by='Prova Objetiva Específica', ascending=False)
    df_nota_final = df_nota_final.drop_duplicates(subset='Aluno', keep='first')
    df_nota_final = df_nota_final.fillna(0)
    df_nota_final = df_nota_final.sort_values(by='Aluno', ascending=True)
    df_nota_final.reset_index(drop=True, inplace=True)
    df_nota_final.to_csv(os.getcwd() + '\\input-files\\CSV_NotasFinais.csv')
    print("Arquivo CSV_NotasFinais.csv gerado com sucesso!")

    # Limpa a memória RAM e mantém só o df_nota_final
    del df_basico, df_especifico, df_discursiva, df_nota_final


def rankear_alunos():
    input_files_path = os.getcwd() + '\\input-files\\'
    files = os.listdir(input_files_path)

    input_ranking = [file for file in files if "CSV_NotasFinais" in file][0]

    df_rank = pd.read_csv(os.getcwd() + '\\input-files\\' + input_ranking)

    notas = df_rank.columns[2:].tolist()
    for nota in notas:
        df_rank = df_rank.sort_values(by = [nota], ascending = False)
        df_rank['Ranking ' + nota] = df_rank[nota].rank(ascending=False, method='dense').astype(int)
        df_rank = df_rank.sort_index()

    df_rank['Texto1'] = "Oi, " + df_rank['Aluno'].str.split().str[0] + "! Tudo bem?\n"
    df_rank['Texto2'] = "Sua Nota Final (objetiva + discursiva): " + df_rank['Nota Final'].astype(str) + "/30.00\n"
    df_rank['Texto3'] = ("Seu ranking da prova discursiva: " + df_rank['Ranking Prova Discursiva'].astype(str) +
                         "ª nota mais alta\n")
    df_rank['Texto4'] = ("Seu ranking da prova Objetiva Básica: " + df_rank['Ranking Prova Objetiva Geral'].astype(str) +
                         "ª nota mais alta\n")
    df_rank['Texto5'] = ("Seu ranking da prova Objetiva Específica: " + df_rank['Ranking Prova Objetiva Específica'].
                         astype(str) + "ª nota mais alta\n")
    df_rank['Texto6'] = ("Seu ranking geral: " + df_rank['Ranking Nota Final'].astype(str) + "ª nota mais alta\n\n")
    df_rank['Texto7'] = "Nas disciplinas:\n"
    df_rank['Texto8'] = ("Língua Portuguesa: " + df_rank['Ranking Língua Portuguesa'].astype(str) +
                         "ª nota mais alta\n")
    df_rank['Texto9'] = ("Noções de direito administrativo e constitucional: " +
                         df_rank['Ranking Noções de Direito Administrativo e Constitucional'].astype(str) +
                         "ª nota mais alta\n")
    df_rank['Texto10'] = ("Noções de raciocínio lógico e matemático: " +
                          df_rank['Ranking Noções de Raciocínio Lógico e Matemático'].astype(str) +
                          "ª nota mais alta\n")
    df_rank['Texto11'] = ("Noções de Informática: " + df_rank['Ranking Noções de Informática'].astype(str) +
                          "ª nota mais alta\n")
    df_rank['Texto12'] = ("Disciplinas do Eixo Transversal: " + df_rank['Ranking Disciplinas do Eixo Transversal'].
                          astype(str) + "ª nota mais alta\n")
    df_rank['Texto13'] = ("Conhecimentos específicos: " + df_rank['Ranking Prova Objetiva Específica'].astype(str) +
                          "ª nota mais alta\n")
    df_rank['Texto14'] = ("Esse simulado teve " + str(len(df_rank)) + " alunos\nContinue firme e bons estudos!!")

    df_rank['Feedback'] = df_rank[
        ['Texto1', 'Texto2', 'Texto3', 'Texto4', 'Texto5', 'Texto6', 'Texto7', 'Texto8', 'Texto9', 'Texto10', 'Texto11',
         'Texto12', 'Texto13', 'Texto14']].apply(lambda x: '\n'.join(x), axis = 1)
    df_rank = df_rank.drop(
        columns = ['Texto1', 'Texto2', 'Texto3', 'Texto4', 'Texto5', 'Texto6', 'Texto7', 'Texto8', 'Texto9', 'Texto10',
                   'Texto11', 'Texto12', 'Texto13', 'Texto14'])

    df_rank['Feedback'] = df_rank['Feedback'].apply(
        lambda x: '\n'.join(line for line in x.splitlines() if line.strip()))
    df_rank['Feedback'] = df_rank['Feedback'].apply(lambda x: '\n'.join(
            line if line != 'Nas disciplinas:' else '\n' + line for line in x.splitlines() if line.strip()))

    df_rank.to_csv(os.getcwd() + '\\input-files\\CSV_Ranking.csv')

    df_rank = df_rank.sort_values(by = 'Ranking Nota Final', ascending = True)
    df_rank.pop('Unnamed: 0')

    df_rank.reset_index(drop = True, inplace = True)
    df_rank.to_csv(os.getcwd() + '\\input-files\\CSV_Ranking.csv')
    print("Arquivo CSV_Ranking.csv gerado com sucesso!")
    try:
        df_rank.to_excel(os.getcwd() + '\\input-files\\Excel_Ranking.xlsx', index=False)
        print("Arquivo Excel_Ranking.xlsx gerado com sucesso!")
    except Exception as e:
        print(f"Aconteceu um erro: {e}")


def arrumar_excel():
    input_files_path = os.getcwd() + '\\input-files\\Excel_Ranking.xlsx'
    output_files_path = os.getcwd() + '\\output-files\\Excel_Ranking.xlsx'

    try:
        workbook = openpyxl.load_workbook(input_files_path)
        sheet = workbook.active

        # 1. Autoadjust column 'A' width
        sheet.column_dimensions['A'].auto_size = True

        # 2. Set cells 'B1' to 'U1' to Wrap Text
        for column in range(ord('B'), ord('U') + 1):
            col_letter = chr(column)
            sheet[f'{col_letter}1'].alignment = Alignment(wrap_text = True)

        # 3. Set cells 'B1' to 'U1' to width=16
        for column in range(ord('B'), ord('U') + 1):
            col_letter = chr(column)
            sheet.column_dimensions[col_letter].width = 16

        # 4. Set column 'V' to width=64
        sheet.column_dimensions['V'].width = 64

        # 5. Set column 'V' to Wrap Text
        sheet['V1'].alignment = Alignment(wrap_text=True)

        # 6. Set row 2 to the last one to center text in cells
        # 7. Set row 2 to the last one to center align in cells
        for row in sheet.iter_rows(min_row = 1, max_row = sheet.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal = "center")

        # 8. Set column 'V' to left align the text
        for cell in sheet['V']:
            cell.alignment = Alignment(horizontal = "left")

        # 9. Set row 1 height=60
        sheet.row_dimensions[1].height = 60

        # 10. Set row 2 to the last one height=225
        for row in range(2, sheet.max_row + 1):
            sheet.row_dimensions[row].height = 225

        # 11. Vertical align all cells in row 1 to the bottom
        sheet.column_dimensions['A'].width = 42
        for cell in sheet[1]:
            cell.alignment = Alignment(vertical = "bottom")

        # 12. Set all cells to Wrap Text
        for row in sheet.iter_rows(min_row = 1, max_row = sheet.max_row, min_col = 1, max_col = sheet.max_column):
            for cell in row:
                cell.alignment = Alignment(wrap_text = True)

        # 13. Put macro
        vba_code = ("""
        Option Explicit

        Private Declare PtrSafe Function OpenClipboard Lib "user32" (ByVal hwnd As LongPtr) As Long
        Private Declare PtrSafe Function CloseClipboard Lib "user32" () As Long
        Private Declare PtrSafe Function EmptyClipboard Lib "user32" () As Long
        Private Declare PtrSafe Function SetClipboardData Lib "user32" (ByVal wFormat As Long, ByVal hMem As LongPtr) As LongPtr
        Private Declare PtrSafe Function GlobalAlloc Lib "kernel32" (ByVal wFlags As Long, ByVal dwBytes As LongPtr) As LongPtr
        Private Declare PtrSafe Function GlobalLock Lib "kernel32" (ByVal hMem As LongPtr) As LongPtr
        Private Declare PtrSafe Function GlobalUnlock Lib "kernel32" (ByVal hMem As LongPtr) As Long
        Private Declare PtrSafe Function GlobalFree Lib "kernel32" (ByVal hMem As LongPtr) As LongPtr
        Private Declare PtrSafe Function lstrcpy Lib "kernel32" (ByVal lpStr1 As Any, ByVal lpStr2 As Any) As LongPtr
        
        Private Const CF_TEXT = 1&
        Private Const GMEM_MOVEABLE = &H2
        
        Sub CopyContent()
        Call StringToClipboard(ActiveCell.Value)
        End Sub
        
        Private Sub StringToClipboard(strText As String)
        Dim lngIdentifier As LongPtr, lngPointer As LongPtr
        lngIdentifier = GlobalAlloc(GMEM_MOVEABLE, Len(strText) + 1)
        lngPointer = GlobalLock(lngIdentifier)
        Call lstrcpy(ByVal lngPointer, strText)
        Call GlobalUnlock(lngIdentifier)
        Call OpenClipboard(0&)
        Call EmptyClipboard
        Call SetClipboardData(CF_TEXT, lngIdentifier)
        Call CloseClipboard
        Call GlobalFree(lngIdentifier)
        End Sub

        """
                    )
        # app = xw.App(visible = True, add_book = False)
        # wb = app.books.open(input_files_path)

        # wb.vba.add_module(vba_code, "MyModule")
        # wb.api.Application.OnKey("^q", "MyModule.MyMacro")
        # wb.save()

        # Save the workbook
        max_row = sheet.max_row
        for row in range(1, max_row + 1):
            cell_value = sheet.cell(row = row, column = 22).value  # Column 'V' is the 22nd column
            if cell_value is not None and isinstance(cell_value, str):
                sheet.cell(row = row, column = 22, value = cell_value.replace('"', ''))

        workbook.save(output_files_path)
        # app.quit()
    except Exception as e:
        print(f"Aconteceu um erro: {e}")


def formatar_excel():
    input_files_path = os.getcwd() + '\\input-files\\Excel_Ranking.xlsx'
    wb = load_workbook(input_files_path)
    ws = wb['Sheet1']

    ws.column_dimensions['V'].width = 63

    for letter in ['A']:
        max_width = 0

        for row_number in range(1, ws.max_row + 1):
            if len(ws[f'{letter}{row_number}'].value) > max_width:
                max_width = len(ws[f'{letter}{row_number}'].value)
            ws['V' + str(row_number)].alignment = Alignment(wrapText = True)
            ws['A' + str(row_number)].alignment = Alignment(vertical = 'center', horizontal = 'left')

        ws.column_dimensions[letter].width = max_width + 1

    for letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U']:
        for row_number in range(1, ws.max_row + 1):
            ws[str(letter) + str(row_number)].alignment = Alignment(vertical = 'center', horizontal = 'center')

    wb.save(input_files_path)

    app = xw.App(visible = True, add_book = False)
    wb = app.books.open(input_files_path)


def colocar_macro():
    input_files_path = os.getcwd() + '\\input-files\\Excel_Ranking.xlsx'

    vba_code = ("""
            Option Explicit

            Private Declare PtrSafe Function OpenClipboard Lib "user32" (ByVal hwnd As LongPtr) As Long
            Private Declare PtrSafe Function CloseClipboard Lib "user32" () As Long
            Private Declare PtrSafe Function EmptyClipboard Lib "user32" () As Long
            Private Declare PtrSafe Function SetClipboardData Lib "user32" (ByVal wFormat As Long, ByVal hMem As LongPtr) As LongPtr
            Private Declare PtrSafe Function GlobalAlloc Lib "kernel32" (ByVal wFlags As Long, ByVal dwBytes As LongPtr) As LongPtr
            Private Declare PtrSafe Function GlobalLock Lib "kernel32" (ByVal hMem As LongPtr) As LongPtr
            Private Declare PtrSafe Function GlobalUnlock Lib "kernel32" (ByVal hMem As LongPtr) As Long
            Private Declare PtrSafe Function GlobalFree Lib "kernel32" (ByVal hMem As LongPtr) As LongPtr
            Private Declare PtrSafe Function lstrcpy Lib "kernel32" (ByVal lpStr1 As Any, ByVal lpStr2 As Any) As LongPtr

            Private Const CF_TEXT = 1&
            Private Const GMEM_MOVEABLE = &H2

            Sub CopyContent()
            Call StringToClipboard(ActiveCell.Value)
            End Sub

            Private Sub StringToClipboard(strText As String)
            Dim lngIdentifier As LongPtr, lngPointer As LongPtr
            lngIdentifier = GlobalAlloc(GMEM_MOVEABLE, Len(strText) + 1)
            lngPointer = GlobalLock(lngIdentifier)
            Call lstrcpy(ByVal lngPointer, strText)
            Call GlobalUnlock(lngIdentifier)
            Call OpenClipboard(0&)
            Call EmptyClipboard
            Call SetClipboardData(CF_TEXT, lngIdentifier)
            Call CloseClipboard
            Call GlobalFree(lngIdentifier)
            End Sub

            """)

    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    workbook = excel.Workbooks.Open(Filename = input_files_path)
    excelModule = workbook.VBProject.VBComponents.Add(1)
    excelModule.CodeModule.AddFromString(vba_code)
    excel.Workbooks(1).Close(SaveChanges = 1)
    excel.Application.Quit()
    del excel

    print(f'Código VBA inserido com sucesso')


