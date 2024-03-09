import pandas as pd
import os
import numpy as np
import openpyxl
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName
import xlwings as xw
import win32com
from win32com.client import Dispatch


def compilar_notas():
    # Dados iniciais ---------------------------------------------------------------------------------------------------
    print('------------------------------------------------')
    print('Dados Iniciais:')
    print('')
    print('Eixos temáticos:')
    eixos_tematicos = {
        "Eixo temático 1": range(1, 11),
        "Eixo temático 2": range(11, 21),
        "Eixo temático 3": range(21, 31),
        "Eixo temático 4": range(31, 41),
        "Eixo temático 5": range(41, 51),
    }
    eixos_tematicos_values = list(eixos_tematicos)
    for i in range(len(eixos_tematicos)):
        print('   ->' + eixos_tematicos_values[i])
    print('')

    q_basico_csv = 0.5  # Quanto vale uma questão básica no .CSV?
    q_especifica_csv = 0.2  # Quanto vale uma questão específica no .CSV?

    print('As questões básicas no CSV valem: ' + str(q_basico_csv))
    print('As questões específicas no CSV valem: ' + str(q_especifica_csv))
    print('')

    print('Pesos das provas:')
    peso_basico = 2

    peso_especifico1 = 2.5
    peso_especifico2 = 7.5
    peso_especifico3 = 2.5
    peso_especifico4 = 2.5
    peso_especifico5 = 10
    pesos_especificos = [peso_especifico1, peso_especifico2, peso_especifico3, peso_especifico4, peso_especifico5]

    print('   ->Questões da prova básica: ' + str(peso_basico))
    for i in range(len(eixos_tematicos)):
        print(f'   ->Questões do Eixo Temático {i + 1}: ' + str(pesos_especificos[i]))
    print('------------------------------------------------')

    # ------------------------------------------------------------------------------------------------------------------

    input_files_path = os.getcwd() + '\\input-files\\'
    files = os.listdir(input_files_path)

    input_basico = [file for file in files if "GER" in file][0]
    input_especifico = [file for file in files if "FIC" in file][0]
    input_discursiva = [file for file in files if "RSI" in file][0]
    df_materias = pd.read_csv(os.getcwd() + '\\materias.csv')

    # Gera o CSV_Basico ------------------------------------------------------------------------------------------------
    df_basico = pd.read_csv(os.getcwd() + '\\input-files\\' + input_basico)
    df_basico.pop('Estado')
    df_basico.pop('Iniciado em')
    df_basico.pop('Completo')
    df_basico.pop('Tempo utilizado')
    df_basico.pop('Avaliar/10,00')
    df_basico = df_basico[~df_basico['Sobrenome'].str.startswith('Média geral')]
    df_basico['Aluno'] = (df_basico['Nome'] + ' ' + df_basico['Sobrenome']).str.title()
    df_basico = df_basico.drop(['Nome', 'Sobrenome'], axis = 1)
    df_basico = df_basico.replace('-', '0,00', regex = True)
    df_basico = df_basico.replace(',', '.', regex = True)
    df_basico = df_basico.apply(pd.to_numeric, errors = 'ignore')
    df_basico = df_basico[['Aluno'] + [col for col in df_basico.columns if col != 'Aluno']]
    df_basico.reset_index(drop = True, inplace = True)

    # Aplica peso 2 nas questões -------------------------------
    # A questão tem que valer 1, no CSV está valendo 0,50, logo:
    # 2 * 0,5 = 1,0
    for i in range(1, 21):
        q_basico_csv = str(q_basico_csv).replace(".", ",")
        coluna_basica = f'Q. {str(i)} /{q_basico_csv}0'
        df_basico[coluna_basica] = df_basico[coluna_basica] * peso_basico

    df_basico['Prova Objetiva Geral'] = np.sum(df_basico.iloc[:, 1:21], axis = 1).round(2)
    for i in range(1, 21):
        df_basico.pop(f'Q. {str(i)} /{q_basico_csv}0')
    df_basico.to_csv(os.getcwd() + '\\output-files\\CSV_Basico.csv')
    print("Arquivo CSV_Basico.csv gerado com sucesso!")

    # Gera o CSV_Especifico --------------------------------------------------------------------------------------------
    df_especifico = pd.read_csv(os.getcwd() + '\\input-files\\' + input_especifico)
    df_especifico.pop('Estado')
    df_especifico.pop('Iniciado em')
    df_especifico.pop('Completo')
    df_especifico.pop('Tempo utilizado')
    df_especifico.pop('Avaliar/10,00')
    df_especifico = df_especifico[~df_especifico['Sobrenome'].str.startswith('Média geral')]
    df_especifico['Aluno'] = (df_especifico['Nome'] + ' ' + df_especifico['Sobrenome']).str.title()
    df_especifico = df_especifico.drop(['Nome', 'Sobrenome'], axis = 1)
    df_especifico = df_especifico.replace('-', '0,00', regex = True)
    df_especifico = df_especifico.replace(',', '.', regex = True)
    df_especifico = df_especifico.apply(pd.to_numeric, errors = 'ignore')
    df_especifico = df_especifico[['Aluno'] + [col for col in df_especifico.columns if col != 'Aluno']]

    # Aplica peso nas questões -------------------------------
    for i in range(1, 51):
        q_especifica_csv = str(q_especifica_csv).replace(".", ",")
        coluna_especifica = f'Q. {str(i)} /{q_especifica_csv}0'
        peso = 1
        if 1 <= i <= 10:
            peso = peso_especifico1
        if 11 <= i <= 20:
            peso = peso_especifico2
        if 21 <= i <= 30:
            peso = peso_especifico3
        if 31 <= i <= 40:
            peso = peso_especifico4
        if 41 <= i <= 50:
            peso = peso_especifico5
        df_especifico[coluna_especifica] = df_especifico[coluna_especifica] * peso

    # Cria a coluna com a soma de todas as questões ----------
    for eixo, colunas in eixos_tematicos.items():
        df_especifico[eixo] = df_especifico.iloc[:, colunas].sum(axis = 1)
    for i in range(1, 51):
        df_especifico.pop(f'Q. {str(i)} /{q_especifica_csv}0')

    df_especifico['Prova Objetiva Específica'] = df_especifico.iloc[:, 1:].sum(axis = 1)
    df_especifico["Prova Objetiva Específica"] = df_especifico["Prova Objetiva Específica"].apply(lambda x: round(x, 2))

    # --------------------------------------------------------
    df_especifico.reset_index(drop = True, inplace = True)
    df_especifico.to_csv(os.getcwd() + '\\output-files\\CSV_Especifico.csv')
    print("Arquivo CSV_Especifico.csv gerado com sucesso!")

    # Gera o CSV_Discursiva --------------------------------------------------------------------------------------------
    df_discursiva = pd.read_csv(os.getcwd() + '\\input-files\\' + input_discursiva)
    df_discursiva.pop('Identificador')
    df_discursiva.pop('Status')
    df_discursiva.pop('Nota máxima')
    df_discursiva.pop('Nota pode ser alterada')
    df_discursiva.pop('Última modificação (envio)')
    df_discursiva.pop('Última modificação (nota)')
    df_discursiva.pop('Comentários de feedback')
    df_discursiva = df_discursiva.dropna(subset = ['Nota'])
    df_discursiva.rename(columns = {'Nome completo': 'Aluno'}, inplace = True)
    df_discursiva['Aluno'] = (df_discursiva['Aluno']).str.title()
    df_discursiva = df_discursiva.replace(',', '.', regex = True)
    df_discursiva = df_discursiva.apply(pd.to_numeric, errors = 'ignore')
    df_discursiva = df_discursiva.loc[df_discursiva.groupby('Aluno')['Nota'].idxmax()]
    df_discursiva.reset_index(drop = True, inplace = True)
    df_discursiva.rename(columns = {'Nota': 'Prova Discursiva'}, inplace = True)
    df_discursiva["Prova Discursiva"] = df_discursiva["Prova Discursiva"].apply(lambda x: round(x, 2))
    df_discursiva.to_csv(os.getcwd() + '\\output-files\\CSV_Discursiva.csv')
    print("Arquivo CSV_Discursiva.csv gerado com sucesso!")

    # Cria o CSV_NotasFinais -------------------------------------------------------------------------------------------
    df_nota_final = pd.merge(df_especifico, df_basico, on = 'Aluno', how = 'outer')
    df_nota_final = pd.merge(df_nota_final, df_discursiva, on = 'Aluno', how = 'outer')
    df_nota_final['Prova Objetiva (Geral + Específica)'] = df_nota_final.iloc[:, 6:8].sum(axis = 1).round(2)
    df_nota_final['Nota Final'] = df_nota_final.iloc[:, 8:10].sum(axis = 1).round(2)
    df_nota_final = df_nota_final.sort_values(by = 'Prova Objetiva Específica', ascending = False)
    df_nota_final = df_nota_final.drop_duplicates(subset = 'Aluno', keep = 'first')
    df_nota_final = df_nota_final.fillna(0)
    df_nota_final = df_nota_final.sort_values(by = 'Aluno', ascending = True)
    df_nota_final.reset_index(drop = True, inplace = True)
    df_nota_final.to_csv(os.getcwd() + '\\output-files\\CSV_NotasFinais.csv')
    print("Arquivo CSV_NotasFinais.csv gerado com sucesso!")

    # Limpa a memória RAM
    del df_basico, df_especifico, df_discursiva, df_nota_final


def rankear_alunos(simulado_num):
    # Dados iniciais ---------------------------------------------------------------------------------------------------
    nota_maxima_basica = 20
    nota_maxima_especifica = 50
    nota_maxima_discursiva = 20

    email_database = 'Usuários (1).xlsx'  # O arquivo deve estar na pasta '\list-email'

    print('------------------------------------------------')
    print('Dados Iniciais:')
    print('')
    print('Nota máxima da prova básica: ' + str(nota_maxima_basica))
    print('Nota máxima da prova específica: ' + str(nota_maxima_especifica))
    print('Nota máxima da prova discursiva: ' + str(nota_maxima_discursiva))
    print('')
    print('------------------------------------------------')
    nota_maxima = nota_maxima_basica + nota_maxima_especifica + nota_maxima_discursiva
    # ------------------------------------------------------------------------------------------------------------------

    input_files_path = os.getcwd() + '\\output-files\\'
    files = os.listdir(input_files_path)

    input_ranking = [file for file in files if "CSV_NotasFinais" in file][0]

    df_rank = pd.read_csv(os.getcwd() + '\\output-files\\' + input_ranking)

    notas = df_rank.columns[2:].tolist()
    for nota in notas:
        df_rank = df_rank.sort_values(by = [nota], ascending = False)
        df_rank['Ranking ' + nota] = df_rank[nota].rank(ascending = False, method = 'dense').astype(int)
        df_rank = df_rank.sort_index()

    df_rank = df_rank.sort_values(by = 'Ranking Nota Final', ascending = True)
    df_rank.reset_index(drop = True, inplace = True)
    df_rank.to_csv(os.getcwd() + '\\output-files\\CSV_Ranking.csv')

    df_rank['Ranking Eixo temático 1'] = (
        df_rank['Ranking Eixo temático 1'].astype(str))
    df_rank['Ranking Eixo temático 2'] = (
        df_rank['Ranking Eixo temático 2'].astype(str))
    df_rank['Ranking Eixo temático 3'] = (
        df_rank['Ranking Eixo temático 3'].astype(str))
    df_rank['Ranking Eixo temático 4'] = (
        df_rank['Ranking Eixo temático 4'].astype(str))
    df_rank['Ranking Eixo temático 5'] = (
        df_rank['Ranking Eixo temático 5'].astype(str))
    df_rank['Ranking Prova Objetiva Geral'] = (
        df_rank['Ranking Prova Objetiva Geral'].astype(str))
    df_rank['Ranking Prova Objetiva Específica'] = (
        df_rank['Ranking Prova Objetiva Específica'].astype(str))
    df_rank['Ranking Prova Discursiva'] = (
        df_rank['Ranking Prova Discursiva'].astype(str))
    df_rank['Ranking Prova Objetiva (Geral + Específica)'] = (
        df_rank['Ranking Prova Objetiva (Geral + Específica)'].astype(str))
    df_rank['Ranking Nota Final'] = (
        df_rank['Ranking Nota Final'].astype(str))

    df_rank.loc[df_rank['Eixo temático 1'] == 0,
    'Ranking Eixo temático 1'] = '-'
    df_rank.loc[df_rank['Eixo temático 2'] == 0,
    'Ranking Eixo temático 2'] = '-'
    df_rank.loc[df_rank['Eixo temático 3'] == 0,
    'Ranking Eixo temático 3'] = '-'
    df_rank.loc[df_rank['Eixo temático 4'] == 0,
    'Ranking Eixo temático 4'] = '-'
    df_rank.loc[df_rank['Eixo temático 5'] == 0,
    'Ranking Eixo temático 5'] = '-'
    df_rank.loc[df_rank['Prova Objetiva Geral'] == 0,
    'Ranking Prova Objetiva Geral'] = '-'
    df_rank.loc[df_rank['Prova Objetiva Específica'] == 0,
    'Ranking Prova Objetiva Específica'] = '-'
    df_rank.loc[df_rank['Prova Discursiva'] == 0,
    'Ranking Prova Discursiva'] = '-'
    df_rank.loc[df_rank['Prova Objetiva (Geral + Específica)'] == 0,
    'Ranking Prova Objetiva (Geral + Específica)'] = '-'
    df_rank.loc[df_rank['Nota Final'] == 0,
    'Ranking Nota Final'] = '-'

    df_rank['Texto1'] = "Oi, " + df_rank['Aluno'].str.split().str[0] + "! Tudo bem?\n"
    df_rank['Texto2'] = ("Sua Nota Final (objetiva + discursiva): " + df_rank['Nota Final'].astype(str) +
                         f"/{str(nota_maxima)}.00\n")
    df_rank['Texto3'] = ("Seu ranking da prova discursiva: " + df_rank['Ranking Prova Discursiva'].astype(str) +
                         "ª nota mais alta\n")
    df_rank['Texto4'] = (
            "Seu ranking da prova Objetiva Básica: " + df_rank['Ranking Prova Objetiva Geral'].astype(str) +
            "ª nota mais alta\n")
    df_rank['Texto5'] = ("Seu ranking da prova Objetiva Específica: " + df_rank['Ranking Prova Objetiva Específica'].
                         astype(str) + "ª nota mais alta\n")
    df_rank['Texto6'] = ("Seu ranking geral: " + df_rank['Ranking Nota Final'].astype(str) + "ª nota mais alta\n\n")
    df_rank['Texto7'] = "Nas disciplinas:\n"
    df_rank['Texto8'] = ("Eixo temático 1: " + df_rank['Ranking Eixo temático 1'].astype(str) +
                         "ª nota mais alta\n")
    df_rank['Texto9'] = ("Eixo temático 2: " +
                         df_rank['Ranking Eixo temático 2'].astype(str) +
                         "ª nota mais alta\n")
    df_rank['Texto10'] = ("Eixo temático 3: " +
                          df_rank['Ranking Eixo temático 3'].astype(str) +
                          "ª nota mais alta\n")
    df_rank['Texto11'] = ("Eixo temático 4: " + df_rank['Ranking Eixo temático 4'].astype(str) +
                          "ª nota mais alta\n")
    df_rank['Texto12'] = ("Eixo temático 5: " + df_rank['Ranking Eixo temático 5'].
                          astype(str) + "ª nota mais alta\n")
    df_rank['Texto13'] = ("Conhecimentos específicos: " + df_rank['Ranking Prova Objetiva Específica'].astype(str) +
                          "ª nota mais alta\n")
    df_rank['Texto14'] = ("Esse simulado teve " + str(len(df_rank)) + " alunos\nContinue firme e bons estudos!!")

    df_rank['Feedback'] = df_rank[
        ['Texto1', 'Texto2', 'Texto3', 'Texto4', 'Texto5', 'Texto6', 'Texto7', 'Texto8', 'Texto9', 'Texto10', 'Texto11',
         'Texto12', 'Texto13', 'Texto14']].apply(lambda x: '\n'.join(x), axis = 1)
    df_rank = df_rank.drop(
            columns = ['Texto1', 'Texto2', 'Texto3', 'Texto4', 'Texto5', 'Texto6', 'Texto7', 'Texto8', 'Texto9',
                       'Texto10',
                       'Texto11', 'Texto12', 'Texto13', 'Texto14'])

    df_rank['Feedback'] = df_rank['Feedback'].apply(
            lambda x: '\n'.join(line for line in x.splitlines() if line.strip()))
    df_rank['Feedback'] = df_rank['Feedback'].apply(lambda x: '\n'.join(
            line if line != 'Nas disciplinas:' else '\n' + line for line in x.splitlines() if line.strip()))

    df_rank['Feedback'] = df_rank['Feedback'].str.replace(' -ª nota mais alta', ' -')

    df_rank = df_rank[df_rank['Ranking Nota Final'] != "-"]
    df_rank.pop('Unnamed: 0')

    df_rank.reset_index(drop = True, inplace = True)

    df_rank.to_csv(os.getcwd() + '\\output-files\\CSV_Ranking.csv')
    print("Arquivo CSV_Ranking.csv gerado com sucesso!")

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------
    email_feedback = []
    for i in range(len(df_rank)):
        var_aluno = df_rank.iat[i, 0].split()[0]
        var1 = df_rank.iat[i, 1].astype(str)  # Nota Eixo temático 1
        var2 = df_rank.iat[i, 2].astype(str)  # Nota Eixo temático 2
        var3 = df_rank.iat[i, 3].astype(str)  # Nota Eixo temático 3
        var4 = df_rank.iat[i, 4].astype(str)  # Nota Eixo temático 4
        var5 = df_rank.iat[i, 5].astype(str)  # Nota Eixo temático 5
        var6 = df_rank.iat[i, 6].astype(str)  # Nota Prova Objetiva Específica
        var7 = df_rank.iat[i, 7].astype(str)  # Nota Prova Objetiva Geral
        var8 = df_rank.iat[i, 8].astype(str)  # Nota Prova Discursiva
        var9 = df_rank.iat[i, 9].astype(str)  # Nota Prova Objetiva (Geral + Específica)
        var10 = df_rank.iat[i, 10].astype(str)  # Nota Final

        var11 = df_rank.iat[i, 11]  # Ranking Eixo temático 1
        var12 = df_rank.iat[i, 12]  # Ranking Eixo temático 2
        var13 = df_rank.iat[i, 13]  # Ranking Eixo temático 3
        var14 = df_rank.iat[i, 14]  # Ranking Eixo temático 4
        var15 = df_rank.iat[i, 15]  # Ranking Eixo temático 5
        var16 = df_rank.iat[i, 16]  # Ranking Prova Objetiva Específica
        var17 = df_rank.iat[i, 17]  # Ranking Prova Objetiva Geral
        var18 = df_rank.iat[i, 18]  # Ranking Prova Discursiva
        var19 = df_rank.iat[i, 19]  # Ranking Prova Objetiva (Geral + Específica)
        var20 = df_rank.iat[i, 20]  # Ranking Final

        filtered_values = df_rank[df_rank['Prova Objetiva Geral'] != 0]['Prova Objetiva Geral']
        var21 = round(filtered_values.mean(), 2)  # Média conhecimentos básicos

        filtered_values = df_rank[df_rank['Prova Objetiva Específica'] != 0]['Prova Objetiva Específica']
        var22 = round(filtered_values.mean(), 2)  # Média conhecimentos específicos

        filtered_values = df_rank[df_rank['Prova Discursiva'] != 0]['Prova Discursiva']
        var23 = round(filtered_values.mean(), 2)  # Média discursiva

        var24 = round(df_rank['Nota Final'].mean(), 2)  # Média Nota Final

        html_string1 = f"<p>Oi, {var_aluno}! Tudo bem? &#128516;</p>"

        if len(str(var6)) == 4:
            var6 = str(var6) + '0'
        if len(str(var7)) == 4:
            var7 = str(var7) + '0'
        if len(str(var8)) == 4:
            var8 = str(var8) + '0'
        if len(str(var10)) == 4:
            var10 = str(var10) + '0'

        if len(str(nota_maxima_basica)) == 2:
            nota_maxima_basica = str(nota_maxima_basica) + '.00'
        if len(str(nota_maxima_especifica)) == 2:
            nota_maxima_especifica = str(nota_maxima_especifica) + '.00'
        if len(str(nota_maxima_discursiva)) == 2:
            nota_maxima_discursiva = str(nota_maxima_discursiva) + '.00'
        if len(str(nota_maxima)) == 2:
            nota_maxima = str(nota_maxima) + '.00'

        if len(str(var21)) == 4:
            var21 = str(var21) + '0'
        if len(str(var22)) == 4:
            var22 = str(var22) + '0'
        if len(str(var23)) == 4:
            var23 = str(var23) + '0'

        html_string1a = "<p><strong>Suas notas:</strong><br/>"
        html_string1b = f"Prova objetiva - Conhecimentos Gerais: <strong>{var7}/{nota_maxima_basica}</strong><br/>"
        html_string1c = f"Prova objetiva - Conhecimentos Específicos: <strong>{var6}/{nota_maxima_especifica}</strong><br/>"
        html_string1d = f"Prova Discursiva: <strong>{var8}/{nota_maxima_discursiva}</strong><br/>"
        html_string3 = f"Sua Nota Final (objetiva + discursiva): <strong>{var10}/{str(nota_maxima)}</strong><br/>"

        html_string2 = "<p><strong>Seu feedback na prova como um todo:</strong><br/>"

        if var18 == '-':
            html_string4 = f"Seu ranking da prova discursiva: <strong>{var18}</strong><br/>"
        if var18 != '-':
            html_string4 = f"Seu ranking da prova discursiva: <strong>{var18}ª</strong> nota mais alta<br/>"

        if var17 == '-':
            html_string5 = f"Seu ranking da prova Objetiva - Conhecimentos Gerais: <strong>{var17}</strong><br/>"
        if var17 != '-':
            html_string5 = f"Seu ranking da prova Objetiva - Conhecimentos Gerais: <strong>{var17}ª</strong> nota mais alta<br/>"

        if var16 == '-':
            html_string6 = f"Seu ranking da prova Objetiva - Conhecimentos Específicos: <strong>{var16}</strong><br/>"
        if var16 != '-':
            html_string6 = f"Seu ranking da prova Objetiva - Conhecimentos Específicos: <strong>{var16}ª</strong> nota mais alta<br/>"

        if var20 == '-':
            html_string7 = f"Seu ranking geral: <strong>{var20}</strong></p>"
        if var20 != '-':
            html_string7 = f"Seu ranking geral: <strong>{var20}ª</strong> nota mais alta</p>"

        html_string8 = "<p><strong>Seu feedback em cada eixo:</strong><br/>"

        if var11 == '-':
            html_string9 = f"Eixo temático 1: <strong>{var11}</strong><br/>"
        if var11 != '-':
            html_string9 = f"Eixo temático 1: <strong>{var11}ª</strong> nota mais alta<br/>"

        if var12 == '-':
            html_string10 = f"Eixo temático 2: <strong>{var12}</strong><br/>"
        if var12 != '-':
            html_string10 = f"Eixo temático 2: <strong>{var12}ª</strong> nota mais alta<br/>"
        if var13 == '-':

            html_string11 = f"Eixo temático 3: <strong>{var13}</strong><br/>"
        if var13 != '-':
            html_string11 = f"Eixo temático 3: <strong>{var13}ª</strong> nota mais alta<br/>"
        if var14 == '-':

            html_string12 = f"Eixo temático 4: <strong>{var14}</strong><br/>"
        if var14 != '-':
            html_string12 = f"Eixo temático 4: <strong>{var14}ª</strong> nota mais alta<br/>"
        if var15 == '-':

            html_string13 = f"Eixo temático 5: <strong>{var15}</strong><br/>"
        if var15 != '-':
            html_string13 = f"Eixo temático 5: <strong>{var15}ª</strong> nota mais alta<br/>"

        html_string14 = f"<p><strong>Esse simulado foi feito por <u>{str(len(df_rank))}</u> alunos</strong><br/>"
        html_string15 = f"Média geral (conhecimentos gerais): <strong>{var21}/{str(nota_maxima_basica)}</strong><br/>"
        html_string16 = f"Média geral (conhecimentos específicos): <strong>{var22}/{str(nota_maxima_especifica)}</strong><br/>"
        html_string17 = f"Média geral (discursiva): <strong>{var23}/{str(nota_maxima_discursiva)}</strong><br/><br/>"
        html_string18 = "Bons estudos!!</p>"
        html_string19 = '<p><font color="gray"><b><i><h1 style="font-size:8pt; ">'
        html_string20 = "Lembrando que esse é um email automático do CONVET!"
        html_string21 = f'</h1></i></b><i><h1 style="font-size:8pt; ">'
        html_string22 = "Se você tentar respondê-lo, ninguém vai ver &#128542;"
        html_string23 = "</h1></i></font></p>"
        html_string24 = f'<img src="https://www.concursosconvet.com.br/assets/img/logotipo/logotipo.png"  width="100" height="100">'

        html_strings = [html_string1, html_string1a, html_string1b, html_string1c, html_string1d, html_string3,
                        html_string2, html_string5, html_string6, html_string4, html_string7, html_string8,
                        html_string9, html_string10, html_string11, html_string12, html_string13, html_string14,
                        html_string15, html_string16, html_string17, html_string18, html_string19, html_string20,
                        html_string21, html_string22, html_string23, html_string24]

        html_string = "".join(html_strings)
        email_feedback.append(html_string)

    df_rank['Email Feedback'] = email_feedback
    df_email_list = pd.read_excel(os.getcwd() + '\\list-email\\' + email_database)

    df_email_list['Aluno'] = (df_email_list['firstname'] + ' ' + df_email_list['lastname']).str.title()

    try:
        df_email_list.pop('id')
        df_email_list.pop('username')
        df_email_list.pop('firstname')
        df_email_list.pop('lastname')
        df_email_list.pop('idnumber')
        df_email_list.pop('institution')
        df_email_list.pop('department')
        df_email_list.pop('phone1')
        df_email_list.pop('phone2')
        df_email_list.pop('city')
        df_email_list.pop('url')
        df_email_list.pop('icq')
        df_email_list.pop('skype')
        df_email_list.pop('aim')
        df_email_list.pop('yahoo')
        df_email_list.pop('msn')
        df_email_list.pop('country')
        df_email_list.pop('profile_field_cpf')
    except Exception as e:
        print(f"Aconteceu um erro: {e}")

    df_email_list = df_email_list[['Aluno', 'email']]

    df_email_list['Aluno'] = df_email_list['Aluno'].str.replace('  ', ' ')
    df_rank['Aluno'] = df_rank['Aluno'].str.replace('  ', ' ')

    df_email_feedback = pd.merge(df_rank, df_email_list, on = 'Aluno', how = 'left')

    try:
        df_email_feedback.pop('Eixo temático 1')
        df_email_feedback.pop('Eixo temático 2')
        df_email_feedback.pop('Eixo temático 3')
        df_email_feedback.pop('Eixo temático 4')
        df_email_feedback.pop('Eixo temático 5')
        df_email_feedback.pop('Prova Objetiva Específica')
        df_email_feedback.pop('Prova Objetiva Geral')
        df_email_feedback.pop('Prova Discursiva')
        df_email_feedback.pop('Prova Objetiva (Geral + Específica)')
        df_email_feedback.pop('Nota Final')
        df_email_feedback.pop('Ranking Eixo temático 1')
        df_email_feedback.pop('Ranking Eixo temático 2')
        df_email_feedback.pop('Ranking Eixo temático 3')
        df_email_feedback.pop('Ranking Eixo temático 4')
        df_email_feedback.pop('Ranking Eixo temático 5')
        df_email_feedback.pop('Ranking Prova Objetiva Específica')
        df_email_feedback.pop('Ranking Prova Objetiva Geral')
        df_email_feedback.pop('Ranking Prova Discursiva')
        df_email_feedback.pop('Ranking Prova Objetiva (Geral + Específica)')
        df_email_feedback.pop('Ranking Nota Final')
        df_email_feedback.pop('Feedback')
    except Exception as e:
        print(e)

    df_email_feedback.to_excel(os.getcwd() + f'\\list-email\\Excel_Email_List - Simulado MAPA {simulado_num}.xlsx', index = False)

    #try:
    #    df_feedback = df_rank[['Aluno', 'Email Feedback']].copy()
    #    df_feedback.to_excel(os.getcwd() + '\\list-email\\Excel_Email_Feedback.xlsx', index = False)
    #    df_rank = df_rank.drop(columns = ['Email Feedback'])
    #    print("Arquivo Excel_Email_Feedback.xlsx gerado com sucesso!")
    #except Exception as e:
    #    print(f"Aconteceu um erro: {e}")

    # ------------------------------------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------------------------------------

    try:
        df_rank.to_excel(os.getcwd() + f'\\output-files\\Excel_Ranking - Simulado MAPA {simulado_num}.xlsx', index = False)
        print('------------------------------------------------')
        print(f'Arquivo Excel_Ranking - Simulado {simulado_num}.xlsx gerado com sucesso!')
        print('')
    except Exception as e:
        print(f'Aconteceu um erro: {e}')


def arrumar_excel(simulado_num):
    input_files_path = os.getcwd() + f'\\output-files\\Excel_Ranking - Simulado MAPA {simulado_num}.xlsx'
    output_files_path = os.getcwd() + f'\\output-files\\Excel_Ranking - Simulado MAPA {simulado_num}.xlsx'

    try:
        workbook = openpyxl.load_workbook(input_files_path)
        sheet = workbook.active

        # 1. Autoadjust column 'A' width
        sheet.column_dimensions['A'].auto_size = True

        # 2. Set cells 'B1' to 'U1' to Wrap Text
        for column in range(ord('B'), ord('U') + 1):
            col_letter = chr(column)
            sheet[f'{col_letter}1'].alignment = Alignment(wrap_text = True)

        # 3. Set cells 'B1' to 'U1' to width=16
        for column in range(ord('B'), ord('U') + 1):
            col_letter = chr(column)
            sheet.column_dimensions[col_letter].width = 16

        # 4. Set column 'V' to width=64
        sheet.column_dimensions['V'].width = 64

        # 5. Set column 'V' to Wrap Text
        sheet['V1'].alignment = Alignment(wrap_text = True)

        # 6. Set row 2 to the last one to center text in cells
        # 7. Set row 2 to the last one to center align in cells
        for row in sheet.iter_rows(min_row = 1, max_row = sheet.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal = "center")

        # 8. Set column 'V' to left align the text
        for cell in sheet['V']:
            cell.alignment = Alignment(horizontal = "left")

        # 9. Set row 1 height=60
        sheet.row_dimensions[1].height = 60

        # 10. Set row 2 to the last one height=225
        for row in range(2, sheet.max_row + 1):
            sheet.row_dimensions[row].height = 225

        # 11. Vertical align all cells in row 1 to the bottom
        sheet.column_dimensions['A'].width = 42
        for cell in sheet[1]:
            cell.alignment = Alignment(vertical = "bottom")

        # 12. Set all cells to Wrap Text
        for row in sheet.iter_rows(min_row = 1, max_row = sheet.max_row, min_col = 1, max_col = sheet.max_column):
            for cell in row:
                cell.alignment = Alignment(wrap_text = True)

        # 14. Put macro
        vba_code = ("""
        Option Explicit

        Private Declare PtrSafe Function OpenClipboard Lib "user32" (ByVal hwnd As LongPtr) As Long
        Private Declare PtrSafe Function CloseClipboard Lib "user32" () As Long
        Private Declare PtrSafe Function EmptyClipboard Lib "user32" () As Long
        Private Declare PtrSafe Function SetClipboardData Lib "user32" (ByVal wFormat As Long, ByVal hMem As LongPtr) As LongPtr
        Private Declare PtrSafe Function GlobalAlloc Lib "kernel32" (ByVal wFlags As Long, ByVal dwBytes As LongPtr) As LongPtr
        Private Declare PtrSafe Function GlobalLock Lib "kernel32" (ByVal hMem As LongPtr) As LongPtr
        Private Declare PtrSafe Function GlobalUnlock Lib "kernel32" (ByVal hMem As LongPtr) As Long
        Private Declare PtrSafe Function GlobalFree Lib "kernel32" (ByVal hMem As LongPtr) As LongPtr
        Private Declare PtrSafe Function lstrcpy Lib "kernel32" (ByVal lpStr1 As Any, ByVal lpStr2 As Any) As LongPtr
        
        Private Const CF_TEXT = 1&
        Private Const GMEM_MOVEABLE = &H2
        
        Sub CopyContent()
        Call StringToClipboard(ActiveCell.Value)
        End Sub
        
        Private Sub StringToClipboard(strText As String)
        Dim lngIdentifier As LongPtr, lngPointer As LongPtr
        lngIdentifier = GlobalAlloc(GMEM_MOVEABLE, Len(strText) + 1)
        lngPointer = GlobalLock(lngIdentifier)
        Call lstrcpy(ByVal lngPointer, strText)
        Call GlobalUnlock(lngIdentifier)
        Call OpenClipboard(0&)
        Call EmptyClipboard
        Call SetClipboardData(CF_TEXT, lngIdentifier)
        Call CloseClipboard
        Call GlobalFree(lngIdentifier)
        End Sub

        """
                    )
        # app = xw.App(visible = True, add_book = False)
        # wb = app.books.open(input_files_path)

        # wb.vba.add_module(vba_code, "MyModule")
        # wb.api.Application.OnKey("^q", "MyModule.MyMacro")
        # wb.save()

        # Save the workbook
        max_row = sheet.max_row
        for row in range(1, max_row + 1):
            cell_value = sheet.cell(row = row, column = 22).value  # Column 'V' is the 22nd column
            if cell_value is not None and isinstance(cell_value, str):
                sheet.cell(row = row, column = 22, value = cell_value.replace('"', ''))

        workbook.save(output_files_path)
        # app.quit()
    except Exception as e:
        print(f"Aconteceu um erro: {e}")


def formatar_excel(simulado_num):
    input_files_path = os.getcwd() + f'\\output-files\\Excel_Ranking - Simulado MAPA {simulado_num}.xlsx'
    wb = load_workbook(input_files_path)
    ws = wb['Sheet1']

    ws.column_dimensions['V'].width = 63

    for letter in ['A']:
        max_width = 0

        for row_number in range(1, ws.max_row + 1):
            if len(ws[f'{letter}{row_number}'].value) > max_width:
                max_width = len(ws[f'{letter}{row_number}'].value)
            ws['V' + str(row_number)].alignment = Alignment(wrapText = True)
            ws['A' + str(row_number)].alignment = Alignment(vertical = 'center', horizontal = 'left')

        ws.column_dimensions[letter].width = max_width + 1

    for letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U']:
        for row_number in range(1, ws.max_row + 1):
            ws[str(letter) + str(row_number)].alignment = Alignment(vertical = 'center', horizontal = 'center')

    wb.save(input_files_path)
    print('------------------------------------------------')
    print("Planilha pronta!")
    print('------------------------------------------------')
    #print("Abrindo Excel")
    #app = xw.App(visible = True, add_book = False)
    #wb = app.books.open(input_files_path)


def colocar_macro(simulado_num):
    input_files_path = os.getcwd() + f'\\output-files\\Excel_Ranking - Simulado {simulado_num}.xlsx'

    vba_code = ("""
            Option Explicit

            Private Declare PtrSafe Function OpenClipboard Lib "user32" (ByVal hwnd As LongPtr) As Long
            Private Declare PtrSafe Function CloseClipboard Lib "user32" () As Long
            Private Declare PtrSafe Function EmptyClipboard Lib "user32" () As Long
            Private Declare PtrSafe Function SetClipboardData Lib "user32" (ByVal wFormat As Long, ByVal hMem As LongPtr) As LongPtr
            Private Declare PtrSafe Function GlobalAlloc Lib "kernel32" (ByVal wFlags As Long, ByVal dwBytes As LongPtr) As LongPtr
            Private Declare PtrSafe Function GlobalLock Lib "kernel32" (ByVal hMem As LongPtr) As LongPtr
            Private Declare PtrSafe Function GlobalUnlock Lib "kernel32" (ByVal hMem As LongPtr) As Long
            Private Declare PtrSafe Function GlobalFree Lib "kernel32" (ByVal hMem As LongPtr) As LongPtr
            Private Declare PtrSafe Function lstrcpy Lib "kernel32" (ByVal lpStr1 As Any, ByVal lpStr2 As Any) As LongPtr

            Private Const CF_TEXT = 1&
            Private Const GMEM_MOVEABLE = &H2

            Sub CopyContent()
            Call StringToClipboard(ActiveCell.Value)
            End Sub

            Private Sub StringToClipboard(strText As String)
            Dim lngIdentifier As LongPtr, lngPointer As LongPtr
            lngIdentifier = GlobalAlloc(GMEM_MOVEABLE, Len(strText) + 1)
            lngPointer = GlobalLock(lngIdentifier)
            Call lstrcpy(ByVal lngPointer, strText)
            Call GlobalUnlock(lngIdentifier)
            Call OpenClipboard(0&)
            Call EmptyClipboard
            Call SetClipboardData(CF_TEXT, lngIdentifier)
            Call CloseClipboard
            Call GlobalFree(lngIdentifier)
            End Sub

            """)


    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    workbook = excel.Workbooks.Open(Filename = input_files_path)
    excelModule = workbook.VBProject.VBComponents.Add(1)
    excelModule.CodeModule.AddFromString(vba_code)
    excel.Workbooks(1).Close(SaveChanges = 1)
    excel.Application.Quit()
    del excel

    print(f'Código VBA inserido com sucesso')
